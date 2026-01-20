import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_pmegp_report(filename="PMEGP_Project_Report.xlsx"):
    wb = openpyxl.Workbook()
    
    # 1. BASIC DATA & FRONT PAGE
    ws_front = wb.active
    ws_front.title = "DPR_FRONT"
    ws_front.append(["PROJECT REPORT UNDER PMEGP SCHEME"])
    ws_front.append([])
    data_info = [
        ["1.0", "Name of the Beneficiary", "SMITA SINGH"],
        ["2.0", "Constitution", "Individual"],
        ["3.0", "District", "UDAIPUR"],
        ["4.0", "Social Category", "General"],
        ["5.0", "Project Cost", 4000000]
    ]
    for row in data_info:
        ws_front.append(row)

    # 2. DEPRECIATION SCHEDULE
    ws_dep = wb.create_sheet("Depreciation")
    ws_dep.append(["STATEMENT SHOWING THE DEPRECIATION ON FIXED ASSETS"])
    ws_dep.append(["Particulars", "Rate", "Value", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"])
    assets = [["Plant & Machinery", 0.15, 3500000], ["Furniture & Fixture", 0.10, 300000]]
    for asset, rate, val in assets:
        row = [asset, f"{rate*100}%", val]
        curr = val
        for _ in range(5):
            dep = curr * rate
            row.append(round(dep, 2))
            curr -= dep
        ws_dep.append(row)

    # 3. REPAYMENT SCHEDULE
    ws_rep = wb.create_sheet("Repayment_Schedule")
    ws_rep.append(["STATEMENT SHOWING THE REPAYMENT OF TERM LOAN & WORKING CAPITAL"])
    ws_rep.append(["Year", "Opening Balance", "Interest (10%)", "Installment", "Closing Balance"])
    loan = 3040000
    inst = loan / 7
    curr_loan = loan
    for y in range(1, 8):
        interest = curr_loan * 0.10
        closing = curr_loan - inst
        ws_rep.append([f"Year {y}", round(curr_loan, 2), round(interest, 2), round(inst, 2), round(max(0, closing), 2)])
        curr_loan = closing

    # 4. CAPACITY & OPERATING EXPENSES
    ws_ops = wb.create_sheet("Operational_Expenses")
    ws_ops.append(["S.No", "Particulars", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"])
    ws_ops.append(["A", "Capacity Utilization (%)", "60%", "70%", "80%", "90%", "100%"])
    
    exp_list = [
        ["1", "Raw materials", 1500000],
        ["2", "Wages", 600000],
        ["3", "Power and Fuel", 120000],
        ["4", "Repairs and Maintenance", 30000],
        ["5", "Administrative Expenses", 50000],
        ["6", "Other Overhead Expenses", 40000]
    ]
    for sn, part, base in exp_list:
        row = [sn, part]
        for y in range(5):
            row.append(round(base * (1 + y*0.1), 2))
        ws_ops.append(row)

    # 5. SALES REALIZATION
    ws_sales = wb.create_sheet("Sales_Realization")
    ws_sales.append(["S.No", "Particulars", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"])
    sales_base = 5000000
    s_row = ["1", "Sales Revenue"]
    for y in range(5):
        s_row.append(round(sales_base * (0.6 + y*0.1), 2))
    ws_sales.append(s_row)

    # Formatting
    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

    wb.save(filename)
    print(f"File {filename} created successfully.")

if __name__ == "__main__":
    generate_pmegp_report()
